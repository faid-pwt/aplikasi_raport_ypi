from PIL import Image
import os
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# --- TAMBAHKAN DUA BARIS INI UNTUK PDF ---
import base64
from weasyprint import HTML

# ==========================================
# 1. PENGATURAN & DAFTAR GURU
# ==========================================
DAFTAR_GURU_KELAS = [
    "Ny. Wasiatul Hasanah",
    "Ny. Laila Afrohah",
    "Ustadzah Nur Hasanah",
    "Ustadz Ali Wafa",
    "Ustadz Moh. Afandi",
    "Ustadz Rosi Efendi",
    "_______________________"
]

# ==========================================
# 2. FUNGSI LOGIKA
# ==========================================
def hitung_predikat(nilai):
    try:
        n = float(nilai)
        if n >= 9: return 'A', 'Sangat Baik'
        elif n >= 7: return 'B', 'Baik'
        elif n >= 6: return 'C', 'Cukup'
        else: return 'D', 'Kurang'
    except:
        return '', ''

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def generate_excel(data_siswa, df_nilai, data_tambahan):
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"
    
    # --- PENGATURAN HALAMAN F4 (216 x 330 mm) ---
    ws.page_setup.paperSize = 14 
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # --- PERBAIKAN 1: OTOMATIS RATA TENGAH ---
    ws.print_options.horizontalCentered = True 
    
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # --- PERBAIKAN 2: MELEBARKAN KOLOM AGAR FULL KERTAS ---
    ws.column_dimensions['A'].width = 5    # Tetap sempit untuk "No"
    ws.column_dimensions['B'].width = 18   # Bantuan Header
    ws.column_dimensions['C'].width = 18   # Bantuan Header
    ws.column_dimensions['D'].width = 10   # Nilai
    ws.column_dimensions['E'].width = 18   # Predikat / Kelas
    ws.column_dimensions['F'].width = 22   # Deskripsi / Semester

    font_bold = Font(bold=True)
    font_bold_u = Font(bold=True, underline="single")
    font_italic = Font(italic=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # --- KOP SURAT & LOGO ---
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    
    try:
        logo = ExcelImage("logo.jpg")
        logo.width, logo.height = 110, 110 
        ws.add_image(logo, 'A1')
    except:
        pass

    font_kop_1 = Font(name='Times New Roman', size=14, bold=True)
    font_kop_2 = Font(name='Times New Roman', size=16, bold=True)
    font_kop_3 = Font(name='Times New Roman', size=14, bold=True)
    font_alamat = Font(name='Times New Roman', size=11, italic=True)

    ws.merge_cells('A1:F1'); ws['A1'] = "MADRASAH TAKMILIYAH ULA"; ws['A1'].font = font_kop_1; ws['A1'].alignment = align_center
    ws.merge_cells('A2:F2'); ws['A2'] = "YPI UMMU CHODIJAH"; ws['A2'].font = font_kop_2; ws['A2'].alignment = align_center
    ws.merge_cells('A3:F3'); ws['A3'] = "KWANYAR BARAT BANGKALAN"; ws['A3'].font = font_kop_3; ws['A3'].alignment = align_center
    
    ws.merge_cells('A4:F4'); ws['A4'] = "Jl. Raya Kwanyar Barat Kecamatan Kwanyar Kabupaten Bangkalan Jawa Timur 69164"
    ws['A4'].font = font_alamat; ws['A4'].alignment = align_center; set_border(ws, 'A4:F4')

    # --- JUDUL & IDENTITAS ---
    ws.merge_cells('A6:F6'); ws['A6'] = "LAPORAN HASIL BELAJAR PESERTA DIDIK"; ws['A6'].font = font_bold; ws['A6'].alignment = align_center
    
    for r in [7, 8, 9]: ws.row_dimensions[r].height = 22 

    ws.merge_cells('A7:B7'); ws['A7'] = "Nama"
    ws.merge_cells('C7:D7'); ws['C7'] = f": {data_siswa['nama']}"
    ws['E7'] = "Kelas"; ws['F7'] = f": {data_siswa['kelas']}"

    ws.merge_cells('A8:B8'); ws['A8'] = "No. Induk"
    ws.merge_cells('C8:D8'); ws['C8'] = f": {data_siswa['no_induk']}"
    ws['E8'] = "Semester"; ws['F8'] = f": {data_siswa['semester']}"

    ws.merge_cells('A9:B9'); ws['A9'] = "No. NSMDT"
    ws.merge_cells('C9:D9'); ws['C9'] = f": {data_siswa['no_statistik']}"
    ws['E9'] = "Tapel"; ws['F9'] = f": {data_siswa['tapel']}"

    # --- TABEL NILAI ---
    row_start = 11
    ws.row_dimensions[row_start].height = 25
    
    ws['A11'] = "No"; ws['A11'].font = font_bold; ws['A11'].alignment = align_center
    ws.merge_cells('B11:C11'); ws['B11'] = "Mata Pelajaran"; ws['B11'].font = font_bold; ws['B11'].alignment = align_center
    ws['D11'] = "Nilai"; ws['D11'].font = font_bold; ws['D11'].alignment = align_center
    ws['E11'] = "Predikat"; ws['E11'].font = font_bold; ws['E11'].alignment = align_center
    ws['F11'] = "Deskripsi"; ws['F11'].font = font_bold; ws['F11'].alignment = align_center

    row_curr = row_start + 1
    for idx, row in df_nilai.iterrows():
        ws.row_dimensions[row_curr].height = 23
        pred, desk = hitung_predikat(row['Nilai'])
        
        ws[f'A{row_curr}'] = idx + 1; ws[f'A{row_curr}'].alignment = align_center
        ws.merge_cells(f'B{row_curr}:C{row_curr}'); ws[f'B{row_curr}'] = row['Mata Pelajaran']
        ws[f'D{row_curr}'] = row['Nilai']; ws[f'D{row_curr}'].alignment = align_center
        ws[f'E{row_curr}'] = pred; ws[f'E{row_curr}'].alignment = align_center
        ws[f'F{row_curr}'] = desk; ws[f'F{row_curr}'].alignment = align_center
        row_curr += 1
    set_border(ws, f'A{row_start}:F{row_curr-1}')

    # --- PERINGKAT, KEPRIBADIAN, KEHADIRAN ---
    ws.row_dimensions[row_curr].height = 25
    ws.merge_cells(f'A{row_curr}:C{row_curr}'); ws[f'A{row_curr}'] = f"Peringkat Ke : {data_tambahan['peringkat']}"
    ws.merge_cells(f'D{row_curr}:F{row_curr}'); ws[f'D{row_curr}'] = f"Dari : {data_tambahan['total_siswa']} Peserta Didik"
    for c in ['A','D']: ws[f'{c}{row_curr}'].alignment = align_center
    set_border(ws, f'A{row_curr}:F{row_curr}')

    r_sub = row_curr + 1
    ws.row_dimensions[r_sub].height = 25
    ws[f'A{r_sub}'] = "No"; ws[f'B{r_sub}'] = "Kepribadian"; ws[f'C{r_sub}'] = "Deskripsi"
    ws[f'D{r_sub}'] = "No"; ws[f'E{r_sub}'] = "Ketidakhadiran"; ws[f'F{r_sub}'] = "Hari"
    for col in ['A','B','C','D','E','F']:
        ws[f'{col}{r_sub}'].font = font_bold; ws[f'{col}{r_sub}'].alignment = align_center

    data_k = [("1", "Kelakuan", data_tambahan['kelakuan']), ("2", "Kerajinan", data_tambahan['kerajinan']), ("3", "Kebersihan", data_tambahan['kebersihan'])]
    data_h = [("1", "Izin", data_tambahan['izin']), ("2", "Sakit", data_tambahan['sakit']), ("3", "Tanpa Keterangan", data_tambahan['alpa'])]
    
    for i in range(3):
        r = r_sub + 1 + i
        ws.row_dimensions[r].height = 22
        ws[f'A{r}'], ws[f'B{r}'], ws[f'C{r}'] = data_k[i]
        ws[f'D{r}'], ws[f'E{r}'], ws[f'F{r}'] = data_h[i]
        for c in ['A','D']: ws[f'{c}{r}'].alignment = align_center
    
    set_border(ws, f'A{r_sub}:C{r_sub+3}'); set_border(ws, f'D{r_sub}:F{r_sub+3}')

    # --- TANDA TANGAN ---
    rt = r_sub + 5
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = f"Kwanyar, {data_tambahan['tanggal']}"; ws[f'E{rt}'].alignment = align_center
    rt += 1
    ws.merge_cells(f'A{rt}:C{rt}'); ws[f'A{rt}'] = "Wali Murid"; ws[f'A{rt}'].alignment = align_center
    
    # --- PERUBAHAN DI SINI (D menjadi E) ---
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = "Guru Kelas"; ws[f'E{rt}'].alignment = align_center
    
    rt += 1
    ws.merge_cells(f'A{rt}:F{rt}'); ws[f'A{rt}'] = "Kepala MDTU Ummu Chodijah"; ws[f'A{rt}'].alignment = align_center
    rt += 4
    ws.merge_cells(f'A{rt}:C{rt}'); ws[f'A{rt}'] = "_______________________"; ws[f'A{rt}'].alignment = align_center
    
    # --- PERUBAHAN DI SINI JUGA (D menjadi E) ---
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = data_tambahan['nama_guru']; ws[f'E{rt}'].font = Font(underline="single"); ws[f'E{rt}'].alignment = align_center
    
    rt += 2
    ws.merge_cells(f'A{rt}:F{rt}'); ws[f'A{rt}'] = "ABDULLAH MUNIR"; ws[f'A{rt}'].font = font_bold_u; ws[f'A{rt}'].alignment = align_center

    # --- PERBAIKAN 3: KUNCI AREA CETAK ---
    ws.print_area = f'A1:F{rt}'

    output = BytesIO()
    wb.save(output)
# --- WATERMARK EXCEL TRANSPARANSI TINGKAT TINGGI ---
    try:
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        # 1. Buka logo
        img_pil = Image.open("logo.png").convert("RGBA")
        datas = img_pil.getdata()
        newData = []
        
        # 2. Proses Pemisahan Warna (Anti menutupi tabel)
        for item in datas:
            if item[3] == 0:
                newData.append((255, 255, 255, 0))
            elif item[0] > 220 and item[1] > 220 and item[2] > 220:
                newData.append((255, 255, 255, 0)) # Buat latar jadi bolong
            else:
                newData.append((item[0], item[1], item[2], 25)) # Transparansi 10%
                
        img_pil.putdata(newData)
        img_pil.save("watermark_temp.png", "PNG")

        # 3. Masukkan ke Excel
        watermark = ExcelImage("watermark_temp.png")
        
        # --- ATUR UKURAN DI SINI ---
        img_w, img_h = 650, 650 
        watermark.width, watermark.height = img_w, img_h

        # --- PENYESUAIAN POSISI AGAR GESER KE KIRI 6MM ---
        posisi_kolom = 0  # Mulai dari Kolom A
        posisi_baris = 12 # Mulai dari Baris 12 (ubah jika kurang naik/turun)
        offset_x = pixels_to_EMU(19) # Jarak 19 piksel dari pinggir kiri
        
        marker = AnchorMarker(col=posisi_kolom, colOff=offset_x, row=posisi_baris, rowOff=0)
        watermark.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(img_w), pixels_to_EMU(img_h)))
        
        ws.add_image(watermark)

    except Exception as e:
        pass

    # --- KUNCI AREA CETAK ---
    ws.print_area = f'A1:F{rt}'

    output = BytesIO()
    wb.save(output)
    
    # Hapus file sementara setelah masuk ke Excel agar server tetap bersih
    try:
        os.remove("watermark_temp.png")
    except:
        pass

    return output.getvalue()
    return output.getvalue()

# ==========================================
# FUNGSI BARU UNTUK GENERATE PDF
# ==========================================
def get_image_base64(path):
    try:
        with open(path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode('utf-8')
    except:
        return ""
def generate_pdf(data_siswa, df_nilai, data_tambahan):
    logo_base64 = get_image_base64("logo.jpg")
    logo_html = f"data:image/jpeg;base64,{logo_base64}" if logo_base64 else ""

    tabel_html = ""
    for idx, row in df_nilai.iterrows():
        pred, desk = hitung_predikat(row['Nilai']) # Memanggil fungsi predikat
        tabel_html += f"""
        <tr>
            <td style="text-align:center;">{idx + 1}</td>
            <td>{row['Mata Pelajaran']}</td>
            <td style="text-align:center;">{row['Nilai']}</td>
            <td style="text-align:center;">{pred}</td>
            <td style="text-align:center;">{desk}</td>
        </tr>
        """
    # --- INI DIA HTML & CSS-NYA (Sudah di dalam Python) ---
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{
                size: 216mm 330mm;
                margin: 15mm 12mm;
                background-color: #ffffff;
            }}
            /* ... (semua kode CSS dan HTML yang saya berikan di pesan sebelumnya) ... */
        </style>
    </head>
    <body>
        </body>
    </html>
    """
    
    pdf_file = "raport_temp.pdf"
    HTML(string=html_content).write_pdf(pdf_file)
    
    with open(pdf_file, "rb") as f:
        return f.read()

# ==========================================
# 3. ANTARMUKA WEB (STREAMLIT)
# ==========================================
st.set_page_config(page_title="Input Raport Digital - Ummu Chodijah", layout="centered")

st.title("📄 Sistem Raport Digital")
st.write("YPI Ummu Chodijah - Kwanyar Barat Bangkalan")

# --- Identitas ---
st.header("1. Identitas Siswa")
c1, c2 = st.columns(2)
with c1:
    nama = st.text_input("Nama Lengkap")
    no_induk = st.text_input("No. Induk")
    no_statistik = st.text_input("No. NSMDT", value="311235260373")
with c2:
    kelas = st.text_input("Kelas", placeholder="I ( Satu ) A")
    semester = st.selectbox("Semester", ["I ( Ganjil )", "II ( Genap )"])
    tapel = st.text_input("Tahun Pelajaran", value="2025-2026")

# --- Form Nilai Mobile Friendly ---
st.header("2. Nilai Mata Pelajaran")
mapel_list = ["Al-Qur'an", "Tajwid", "Tafsir", "Fiqih", "Tauhid", "Hadits", "Akhlaq", "Bahasa Arab", 
              "Bahasa Madura Halus", "Nahwu", "Shorof", "Tarikh Islam", "Tahsinul Khot", "Qiroatul Khot", 
              "I'rob", "I'lal", "Imla'", "Muhafadzoh", "Aswaja", "Praktek Sholat", "Do'a"]

pilihan_nilai = [""] + [str(i) for i in range(11)]

# Kita bagi menjadi 2 kolom. 
# Di Laptop akan jadi kiri-kanan, di HP otomatis bersusun ke bawah.
col1, col2 = st.columns(2)

nilai_siswa = [] # Tempat menyimpan hasil input

for i, mapel in enumerate(mapel_list):
    # Memasukkan ke kolom kiri dan kanan secara bergantian
    if i % 2 == 0:
        with col1:
            nilai = st.selectbox(mapel, options=pilihan_nilai, key=f"mapel_{i}")
    else:
        with col2:
            nilai = st.selectbox(mapel, options=pilihan_nilai, key=f"mapel_{i}")
    
    # Simpan nilai yang dipilih
    nilai_siswa.append(nilai)

# Ubah kembali data yang diinput menjadi DataFrame 
# (Sangat penting agar kode cetak PDF dan Excel di bawahnya tidak error)
edited_df = pd.DataFrame({
    "Mata Pelajaran": mapel_list,
    "Nilai": nilai_siswa
})

# --- Tambahan ---
st.header("3. Peringkat, Kehadiran & Guru")
ca, cb, cc = st.columns(3)
with ca:
    peringkat = st.text_input("Peringkat Ke")
    total_siswa = st.text_input("Dari Total Siswa", value="1")
with cb:
    kelakuan = st.selectbox("Kelakuan", ["Baik", "Cukup", "Kurang"])
    kerajinan = st.selectbox("Kerajinan", ["Baik", "Cukup", "Kurang"])
    kebersihan = st.selectbox("Kebersihan", ["Baik", "Cukup", "Kurang"])
with cc:
    izin = st.text_input("Izin (Hari)", value="0")
    sakit = st.text_input("Sakit (Hari)", value="0")
    alpa = st.text_input("Tanpa Keterangan (Hari)", value="0")

tanggal = st.text_input("Tanggal Raport", value="30 April 2026")
guru_terpilih = st.selectbox("Pilih Nama Guru Kelas", DAFTAR_GURU_KELAS)

# --- Tombol Eksekusi ---
if st.button("🚀 Buat Raport Excel (Klik Disini...)"):
    if not nama:
        st.error("Nama Siswa tidak boleh kosong!")
    else:
        data_siswa = {'nama': nama, 'no_induk': no_induk, 'no_statistik': no_statistik, 'kelas': kelas, 'semester': semester, 'tapel': tapel}
        data_tambahan = {
            'peringkat': peringkat, 'total_siswa': total_siswa, 
            'kelakuan': kelakuan, 'kerajinan': kerajinan, 'kebersihan': kebersihan,
            'izin': izin, 'sakit': sakit, 'alpa': alpa, 
            'tanggal': tanggal, 'nama_guru': guru_terpilih
        }
        
        excel_data = generate_excel(data_siswa, edited_df, data_tambahan)
        
        st.success(f"Raport {nama} siap dicetak!")
        st.download_button(
            label="⬇️ Klik untuk Download Excel",
            data=excel_data,
            file_name=f"Raport_{nama.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    # --- Tombol Eksekusi ---
col_btn1, col_btn2 = st.columns(2)

# with col_btn1:
#     if st.button("🚀 Buat Raport Excel"):
#         if not nama:
#             st.error("Nama Siswa tidak boleh kosong!")
#         else:
#             data_siswa = {'nama': nama, 'no_induk': no_induk, 'no_statistik': no_statistik, 'kelas': kelas, 'semester': semester, 'tapel': tapel}
#             data_tambahan = {'peringkat': peringkat, 'total_siswa': total_siswa, 'kelakuan': kelakuan, 'kerajinan': kerajinan, 'kebersihan': kebersihan, 'izin': izin, 'sakit': sakit, 'alpa': alpa, 'tanggal': tanggal, 'nama_guru': guru_terpilih}
            
#             excel_data = generate_excel(data_siswa, edited_df, data_tambahan)
#             st.success(f"Excel siap!")
#             st.download_button("⬇️ Download Excel", data=excel_data, file_name=f"Raport_{nama}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# with col_btn2:
#     if st.button("📄 Buat Raport PDF"):
#         if not nama:
#             st.error("Nama Siswa tidak boleh kosong!")
#         else:
#             data_siswa = {'nama': nama, 'no_induk': no_induk, 'no_statistik': no_statistik, 'kelas': kelas, 'semester': semester, 'tapel': tapel}
#             data_tambahan = {'peringkat': peringkat, 'total_siswa': total_siswa, 'kelakuan': kelakuan, 'kerajinan': kerajinan, 'kebersihan': kebersihan, 'izin': izin, 'sakit': sakit, 'alpa': alpa, 'tanggal': tanggal, 'nama_guru': guru_terpilih}
            
#             pdf_bytes = generate_pdf(data_siswa, edited_df, data_tambahan)
#             st.success(f"PDF siap!")
#             st.download_button("⬇️ Download PDF", data=pdf_bytes, file_name=f"Raport_{nama}.pdf", mime="application/pdf")
    
        

        