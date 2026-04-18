import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

# ==========================================
# 1. PENGATURAN & DAFTAR GURU
# ==========================================
DAFTAR_GURU_KELAS = [
    "Ny. Wasiatul Hasanah",
    "Ust. Ahmad Fauzi, S.Pd.I",
    "Ustzh. Siti Aminah, S.Pd",
    "_______________________"
]

# ==========================================
# 2. FUNGSI LOGIKA
# ==========================================
def hitung_predikat(nilai):
    try:
        n = float(nilai)
        if n >= 9: return 'A', 'Sangat Baik'
        elif n >= 7: return 'B', 'Baik'
        elif n >= 6: return 'C', 'Cukup'
        else: return 'D', 'Kurang'
    except:
        return '', ''

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def generate_excel(data_siswa, df_nilai, data_tambahan):
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"
    
    # --- PENGATURAN HALAMAN F4 (216 x 330 mm) ---
    ws.page_setup.paperSize = 14 
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # Mengatur Margin 
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Pengaturan Lebar Kolom
    ws.column_dimensions['A'].width = 14 
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

    font_bold = Font(bold=True)
    font_bold_u = Font(bold=True, underline="single")
    font_italic = Font(italic=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # --- KOP SURAT & LOGO ---
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 26
    
    try:
        logo = ExcelImage("logo.jpg")
        logo.width, logo.height = 75, 75 
        ws.add_image(logo, 'A1')
    except:
        pass

    ws.merge_cells('A1:F1'); ws['A1'] = "MADRASAH DINIYAH ULA"; ws['A1'].font = font_bold; ws['A1'].alignment = align_center
    ws.merge_cells('A2:F2'); ws['A2'] = "YPI UMMU CHODIJAH"; ws['A2'].font = font_bold; ws['A2'].alignment = align_center
    ws.merge_cells('A3:F3'); ws['A3'] = "KWANYAR BARAT BANGKALAN"; ws['A3'].font = font_bold; ws['A3'].alignment = align_center
    
    ws.merge_cells('A4:F4'); ws['A4'] = "Jl. Raya Kwanyar Barat Kecamatan Kwanyar Kabupaten Bangkalan Jawa Timur 69164"
    ws['A4'].font = font_italic; ws['A4'].alignment = align_center; set_border(ws, 'A4:F4')

    # --- JUDUL & IDENTITAS ---
    ws.merge_cells('A6:F6'); ws['A6'] = "LAPORAN HASIL BELAJAR PESERTA DIDIK"; ws['A6'].font = font_bold; ws['A6'].alignment = align_center
    
    # Merenggangkan baris identitas
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20

    ws['A7'] = "Nama"; ws['B7'] = f": {data_siswa['nama']}"; ws['D7'] = "Kelas"; ws['E7'] = f": {data_siswa['kelas']}"
    ws['A8'] = "No. Induk"; ws['B8'] = f": {data_siswa['no_induk']}"; ws['D8'] = "Semester"; ws['E8'] = f": {data_siswa['semester']}"
    ws['A9'] = "No. Statistik"; ws['B9'] = f": {data_siswa['no_statistik']}"; ws['D9'] = "Tapel"; ws['E9'] = f": {data_siswa['tapel']}"

    # --- TABEL NILAI ---
    row_start = 11
    ws.row_dimensions[row_start].height = 25 # Header tabel lebih tinggi
    headers = ["No", "Mata Pelajaran", "Nilai", "Predikat", "Deskripsi"]
    for col, val in zip(['A','B','C','D','E'], headers):
        ws[f'{col}{row_start}'] = val; ws[f'{col}{row_start}'].font = font_bold; ws[f'{col}{row_start}'].alignment = align_center
    ws.merge_cells(f'E{row_start}:F{row_start}')

    row_curr = row_start + 1
    for idx, row in df_nilai.iterrows():
        ws.row_dimensions[row_curr].height = 23 # Merenggangkan SETIAP baris mata pelajaran
        
        pred, desk = hitung_predikat(row['Nilai'])
        ws[f'A{row_curr}'] = idx + 1; ws[f'A{row_curr}'].alignment = align_center
        ws[f'B{row_curr}'] = row['Mata Pelajaran']
        ws[f'C{row_curr}'] = row['Nilai']; ws[f'C{row_curr}'].alignment = align_center
        ws[f'D{row_curr}'] = pred; ws[f'D{row_curr}'].alignment = align_center
        ws[f'E{row_curr}'] = desk; ws[f'E{row_curr}'].alignment = align_center
        ws.merge_cells(f'E{row_curr}:F{row_curr}')
        row_curr += 1
    set_border(ws, f'A{row_start}:F{row_curr-1}')

    # --- PERINGKAT ---
    ws.row_dimensions[row_curr].height = 25 # Baris peringkat dilebarkan
    ws.merge_cells(f'A{row_curr}:C{row_curr}'); ws[f'A{row_curr}'] = f"Peringkat Ke : {data_tambahan['peringkat']}"; ws[f'A{row_curr}'].alignment = align_center
    ws.merge_cells(f'D{row_curr}:F{row_curr}'); ws[f'D{row_curr}'] = f"Dari : {data_tambahan['total_siswa']} Peserta Didik"; ws[f'D{row_curr}'].alignment = align_center
    row_current = row_curr
    set_border(ws, f'A{row_current}:F{row_current}')

    # --- KEPRIBADIAN & KEHADIRAN ---
    r_sub = row_current + 1
    ws.row_dimensions[r_sub].height = 25 # Header kepribadian dilebarkan
    ws[f'A{r_sub}'] = "No"; ws[f'B{r_sub}'] = "Kepribadian"; ws[f'C{r_sub}'] = "Deskripsi"
    ws[f'D{r_sub}'] = "No"; ws[f'E{r_sub}'] = "Ketidakhadiran"; ws[f'F{r_sub}'] = "Hari"
    for col in ['A','B','C','D','E','F']:
        ws[f'{col}{r_sub}'].font = font_bold; ws[f'{col}{r_sub}'].alignment = align_center

    data_k = [("1", "Kelakuan", data_tambahan['kelakuan']), ("2", "Kerajinan", data_tambahan['kerajinan']), ("3", "Kebersihan", data_tambahan['kebersihan'])]
    data_h = [("1", "Izin", data_tambahan['izin']), ("2", "Sakit", data_tambahan['sakit']), ("3", "Tanpa Keterangan", data_tambahan['alpa'])]
    
    for i in range(3):
        r = r_sub + 1 + i
        ws.row_dimensions[r].height = 22 # Baris data kepribadian dilebarkan
        ws[f'A{r}'], ws[f'B{r}'], ws[f'C{r}'] = data_k[i]
        ws[f'D{r}'], ws[f'E{r}'], ws[f'F{r}'] = data_h[i]
        for c in ['A','D']: ws[f'{c}{r}'].alignment = align_center
    
    set_border(ws, f'A{r_sub}:C{r_sub+3}'); set_border(ws, f'D{r_sub}:F{r_sub+3}')

    # --- TANDA TANGAN ---
    rt = r_sub + 5
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = f"Kwanyar, {data_tambahan['tanggal']}"; ws[f'E{rt}'].alignment = align_center
    rt += 1
    ws.merge_cells(f'A{rt}:B{rt}'); ws[f'A{rt}'] = "Wali Murid"; ws[f'A{rt}'].alignment = align_center
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = "Guru Kelas"; ws[f'E{rt}'].alignment = align_center
    rt += 1
    ws.merge_cells(f'A{rt}:F{rt}'); ws[f'A{rt}'] = "Kepala Madrasah Ummu Chodijah"; ws[f'A{rt}'].alignment = align_center
    rt += 4
    ws.merge_cells(f'A{rt}:B{rt}'); ws[f'A{rt}'] = "_______________________"; ws[f'A{rt}'].alignment = align_center
    ws.merge_cells(f'E{rt}:F{rt}'); ws[f'E{rt}'] = data_tambahan['nama_guru']; ws[f'E{rt}'].font = Font(underline="single"); ws[f'E{rt}'].alignment = align_center
    rt += 2
    ws.merge_cells(f'A{rt}:F{rt}'); ws[f'A{rt}'] = "ABDULLAH MUNIR, S.Pd.I"; ws[f'A{rt}'].font = font_bold_u; ws[f'A{rt}'].alignment = align_center

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 3. ANTARMUKA WEB (STREAMLIT)
# ==========================================
st.set_page_config(page_title="Input Raport Digital - Ummu Chodijah", layout="centered")

st.title("📄 Sistem Raport Digital")
st.write("YPI Ummu Chodijah - Kwanyar Barat Bangkalan")

# --- Identitas ---
st.header("1. Identitas Siswa")
c1, c2 = st.columns(2)
with c1:
    nama = st.text_input("Nama Lengkap")
    no_induk = st.text_input("No. Induk")
    no_statistik = st.text_input("No. Statistik", value="311 2 1520 0737")
with c2:
    kelas = st.text_input("Kelas", placeholder="I ( Satu ) A")
    semester = st.selectbox("Semester", ["I ( Ganjil )", "II ( Genap )"])
    tapel = st.text_input("Tahun Pelajaran", value="2025-2026")

# --- Tabel Nilai Dropdown ---
st.header("2. Nilai Mata Pelajaran")
mapel_list = ["Al-Qur'an", "Tajwid", "Tafsir", "Fiqih", "Tauhid", "Hadits", "Akhlaq", "Bahasa Arab", 
              "Bahasa Madura Halus", "Nahwu", "Shorof", "Tarikh Islam", "Tahsinul Khot", "Qiroatul Khot", 
              "I'rob", "I'lal", "Imla'", "Muhafadzoh", "Aswaja", "Praktek Sholat", "Do'a"]
df_init = pd.DataFrame({"Mata Pelajaran": mapel_list, "Nilai": ["" for _ in mapel_list]})
pilihan_nilai = [""] + [str(i) for i in range(11)]

edited_df = st.data_editor(
    df_init, num_rows="fixed", use_container_width=True, hide_index=True,
    column_config={
        "Nilai": st.column_config.SelectboxColumn("Nilai", options=pilihan_nilai, width="medium")
    }
)

# --- Tambahan ---
st.header("3. Peringkat, Kehadiran & Guru")
ca, cb, cc = st.columns(3)
with ca:
    peringkat = st.text_input("Peringkat Ke")
    total_siswa = st.text_input("Dari Total Siswa", value="1")
with cb:
    kelakuan = st.selectbox("Kelakuan", ["Baik", "Cukup", "Kurang"])
    kerajinan = st.selectbox("Kerajinan", ["Baik", "Cukup", "Kurang"])
    kebersihan = st.selectbox("Kebersihan", ["Baik", "Cukup", "Kurang"])
with cc:
    izin = st.text_input("Izin (Hari)", value="0")
    sakit = st.text_input("Sakit (Hari)", value="0")
    alpa = st.text_input("Tanpa Keterangan (Hari)", value="0")

tanggal = st.text_input("Tanggal Raport", value="30 April 2026")
guru_terpilih = st.selectbox("Pilih Nama Guru Kelas", DAFTAR_GURU_KELAS)

# --- Tombol Eksekusi ---
if st.button("🚀 Buat Raport Excel (F4 Ready)"):
    if not nama:
        st.error("Nama Siswa tidak boleh kosong!")
    else:
        data_siswa = {'nama': nama, 'no_induk': no_induk, 'no_statistik': no_statistik, 'kelas': kelas, 'semester': semester, 'tapel': tapel}
        data_tambahan = {
            'peringkat': peringkat, 'total_siswa': total_siswa, 
            'kelakuan': kelakuan, 'kerajinan': kerajinan, 'kebersihan': kebersihan,
            'izin': izin, 'sakit': sakit, 'alpa': alpa, 
            'tanggal': tanggal, 'nama_guru': guru_terpilih
        }
        
        excel_data = generate_excel(data_siswa, edited_df, data_tambahan)
        
        st.success(f"Raport {nama} siap dicetak!")
        st.download_button(
            label="⬇️ Klik untuk Download Excel",
            data=excel_data,
            file_name=f"Raport_{nama.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )