import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

def hitung_predikat(nilai):
    try:
        n = float(nilai)
        if n >= 9: return 'A', 'Sangat Baik'
        elif n >= 7: return 'B', 'Baik'
        elif n >= 6: return 'C', 'Cukup'
        else: return 'D', 'Kurang'
    except ValueError:
        return '', ''

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border

def main():
    print("=====================================================")
    print("    SISTEM INPUT RAPORT YPI UMMU CHODIJAH            ")
    print("=====================================================\n")

    # 1. Input Data
    print("--- 1. Masukkan Data Siswa ---")
    nama = input("Nama Siswa        : ")
    no_induk = input("No. Induk         : ")
    no_statistik = input("No. Statistik     : ")
    kelas = input("Kelas             : ")
    semester = input("Semester          : ")
    tapel = input("Tahun Pelajaran   : ")

    print("\n--- 2. Masukkan Nilai Mata Pelajaran ---")
    daftar_mapel = [
        "Al-Qur'an", "Tajwid", "Tafsir", "Fiqih", "Tauhid", "Hadits", 
        "Akhlaq", "Bahasa Arab", "Bahasa Madura Halus", "Nahwu", 
        "Shorof", "Tarikh Islam", "Tahsinul Khot", "Qiroatul Khot", 
        "I'rob", "I'lal", "Imla'", "Muhafadzoh", "Aswaja", 
        "Praktek Sholat", "Do'a"
    ]
    
    data_nilai = []
    for i, mapel in enumerate(daftar_mapel, start=1):
        nilai_input = input(f"Nilai {mapel} (Kosongkan jika tidak ada): ")
        if nilai_input.strip() == "":
            nilai, predikat, deskripsi = "", "", ""
        else:
            nilai = nilai_input
            predikat, deskripsi = hitung_predikat(nilai)
        data_nilai.append((i, mapel, nilai, predikat, deskripsi))

    print("\n--- 3. Masukkan Data Tambahan ---")
    peringkat = input("Peringkat Ke               : ")
    total_siswa = input("Dari Total Siswa           : ")
    
    print("\n[Kepribadian]")
    kelakuan = input("Kelakuan                   : ")
    kerajinan = input("Kerajinan                  : ")
    kebersihan = input("Kebersihan                 : ")

    print("\n[Ketidakhadiran]")
    izin = input("Izin                       : ")
    sakit = input("Sakit                      : ")
    alpa = input("Tanpa Keterangan           : ")

    tgl_default = "31 Maret 2022"
    tanggal_raport = input(f"\nTanggal Raport (Enter untuk default '{tgl_default}'): ")
    if not tanggal_raport: tanggal_raport = tgl_default

    # 2. Pembuatan File Excel dengan Layout Khusus
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"

    # --- PENGATURAN UKURAN KERTAS F4 (FOLIO) ---
    ws.page_setup.paperSize = 14

    # Pengaturan Lebar Kolom
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    # Pengaturan Tinggi Baris Kop Surat agar Logo Muat
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # Gaya Font dan Alignment
    font_bold = Font(bold=True)
    font_bold_underline = Font(bold=True, underline="single")
    font_italic = Font(italic=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # --- KOP SURAT & LOGO ---
    try:
        logo = ExcelImage("logo.jpg")
        logo.width = 80   
        logo.height = 80  
        ws.add_image(logo, 'A1') 
    except FileNotFoundError:
        print("\n[WARNING] File 'logo.jpg' tidak ditemukan. Melanjutkan pembuatan raport tanpa logo.")

    ws.merge_cells('B1:F1')
    ws['B1'] = "MADRASAH DINIYAH ULA"
    ws['B1'].font = font_bold
    ws['B1'].alignment = align_center

    ws.merge_cells('B2:F2')
    ws['B2'] = "YPI UMMU CHODIJAH"
    ws['B2'].font = font_bold
    ws['B2'].alignment = align_center

    ws.merge_cells('B3:F3')
    ws['B3'] = "KWANYAR BANGKALAN"
    ws['B3'].font = font_bold
    ws['B3'].alignment = align_center

    ws.merge_cells('A4:F4')
    ws['A4'] = "Jl. Raya Kwanyar Barat Kecamatan Kwanyar Kabupaten Bangkalan Jawa Timur 69164"
    ws['A4'].font = font_italic
    ws['A4'].alignment = align_center
    set_border(ws, 'A4:F4')

    # --- JUDUL DOKUMEN ---
    ws.merge_cells('A6:F6')
    ws['A6'] = "LAPORAN HASIL BELAJAR PESERTA DIDIK"
    ws['A6'].font = font_bold
    ws['A6'].alignment = align_center

    # --- IDENTITAS SISWA ---
    ws['A7'] = "Nama"
    ws['B7'] = f": {nama}"
    ws['D7'] = "Kelas"
    ws['E7'] = f": {kelas}"

    ws['A8'] = "No. Induk"
    ws['B8'] = f": {no_induk}"
    ws['D8'] = "Semester"
    ws['E8'] = f": {semester}"

    ws['A9'] = "No. Statistik"
    ws['B9'] = f": {no_statistik}"
    ws['D9'] = "Tapel"
    ws['E9'] = f": {tapel}"

    # --- TABEL UTAMA ---
    row_start = 11
    headers = ["No", "Mata Pelajaran", "Nilai", "Predikat", "Deskripsi"]
    
    for col, value in zip(['A', 'B', 'C', 'D', 'E'], headers):
        cell = ws[f'{col}{row_start}']
        cell.value = value
        cell.font = font_bold
        cell.alignment = align_center
    
    ws.merge_cells(f'E{row_start}:F{row_start}')
    
    row_current = row_start + 1
    for no, mapel, nilai, predikat, deskripsi in data_nilai:
        ws[f'A{row_current}'] = no
        ws[f'A{row_current}'].alignment = align_center
        ws[f'B{row_current}'] = mapel
        ws[f'C{row_current}'] = nilai
        ws[f'C{row_current}'].alignment = align_center
        ws[f'D{row_current}'] = predikat
        ws[f'D{row_current}'].alignment = align_center
        ws[f'E{row_current}'] = deskripsi
        ws[f'E{row_current}'].alignment = align_center
        
        ws.merge_cells(f'E{row_current}:F{row_current}')
        row_current += 1

    set_border(ws, f'A{row_start}:F{row_current-1}')

    # --- PERINGKAT ---
    ws.merge_cells(f'A{row_current}:C{row_current}')
    ws[f'A{row_current}'] = f"Peringkat Ke :               {peringkat}"
    ws[f'A{row_current}'].alignment = align_center

    ws.merge_cells(f'D{row_current}:F{row_current}')
    ws[f'D{row_current}'] = f"Dari :      {total_siswa}      Peserta Didik"
    ws[f'D{row_current}'].alignment = align_center
    
    set_border(ws, f'A{row_current}:F{row_current}')

    # --- TABEL KEPRIBADIAN & KEHADIRAN ---
    row_sub = row_current + 1
    
    ws[f'A{row_sub}'] = "No"; ws[f'B{row_sub}'] = "Kepribadian"; ws[f'C{row_sub}'] = "Deskripsi"
    ws[f'D{row_sub}'] = "No"; ws[f'E{row_sub}'] = "Ketidakhadiran"; ws[f'F{row_sub}'] = "Hari"
    for col in ['A','B','C','D','E','F']:
        ws[f'{col}{row_sub}'].font = font_bold
        ws[f'{col}{row_sub}'].alignment = align_center

    kepribadian_data = [("1", "Kelakuan", kelakuan), ("2", "Kerajinan", kerajinan), ("3", "Kebersihan", kebersihan)]
    kehadiran_data = [("1", "Izin", izin), ("2", "Sakit", sakit), ("3", "Tanpa Keterangan", alpa)]

    for i in range(3):
        r = row_sub + 1 + i
        ws[f'A{r}'] = kepribadian_data[i][0]; ws[f'A{r}'].alignment = align_center
        ws[f'B{r}'] = kepribadian_data[i][1]
        ws[f'C{r}'] = kepribadian_data[i][2]
        
        ws[f'D{r}'] = kehadiran_data[i][0]; ws[f'D{r}'].alignment = align_center
        ws[f'E{r}'] = kehadiran_data[i][1]
        ws[f'F{r}'] = kehadiran_data[i][2]

    set_border(ws, f'A{row_sub}:C{row_sub+3}')
    set_border(ws, f'D{row_sub}:F{row_sub+3}')

    # --- TANDA TANGAN ---
    row_ttd = row_sub + 5 
    
    ws.merge_cells(f'E{row_ttd}:F{row_ttd}')
    ws[f'E{row_ttd}'] = f"Kwanyar, {tanggal_raport}"
    ws[f'E{row_ttd}'].alignment = align_center

    row_ttd += 1
    ws.merge_cells(f'A{row_ttd}:B{row_ttd}')
    ws[f'A{row_ttd}'] = "Wali Murid"
    ws[f'A{row_ttd}'].alignment = align_center

    ws.merge_cells(f'E{row_ttd}:F{row_ttd}')
    ws[f'E{row_ttd}'] = "Guru Kelas"
    ws[f'E{row_ttd}'].alignment = align_center
    
    row_ttd += 1
    ws.merge_cells(f'A{row_ttd}:F{row_ttd}') 
    ws[f'A{row_ttd}'] = "Kepala Madrasah Ummu Chodijah"
    ws[f'A{row_ttd}'].alignment = align_center

    row_ttd += 4
    ws.merge_cells(f'A{row_ttd}:B{row_ttd}')
    ws[f'A{row_ttd}'] = "_______________________"
    ws[f'A{row_ttd}'].alignment = align_center

    ws.merge_cells(f'E{row_ttd}:F{row_ttd}')
    ws[f'E{row_ttd}'] = "Ny. Wasiatul Hasanah"
    ws[f'E{row_ttd}'].font = Font(underline="single")
    ws[f'E{row_ttd}'].alignment = align_center

    row_ttd += 2
    ws.merge_cells(f'A{row_ttd}:F{row_ttd}')
    ws[f'A{row_ttd}'] = "ABDULLAH MUNIR, S.Pd.I"
    ws[f'A{row_ttd}'].font = font_bold_underline
    ws[f'A{row_ttd}'].alignment = align_center

    # Ekspor
    nama_file = f"Raport_{nama.replace(' ', '_')}_{kelas}.xlsx"
    wb.save(nama_file)
    print(f"\n[SUKSES] Layout raport dengan ukuran A4 berhasil disimpan di: {nama_file}")

if __name__ == "__main__":
    main()